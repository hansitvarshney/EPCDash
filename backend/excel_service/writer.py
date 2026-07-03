import os
from datetime import datetime
from typing import Any, Dict, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from backend.excel_service.config import WorkbookTemplateConfig, SheetMapping
from backend.excel_service.registry import TemplateRegistry
from backend.excel_service.styles import SEVERITY_STYLES, THIN_BORDER
from backend.excel_service.bootstrap import ensure_template_exists


class ExcelWriterService:
    """
    Thin, config-driven abstraction over an openpyxl workbook.

    Node/application code interacts only through logical `(category, field)`
    keys -- never raw sheet names or cell references. Swapping to a real
    client's live workbook later is purely a `TemplateRegistry` profile
    change (see `backend/excel_service/config.py`).
    """

    def __init__(self, profile: Optional[str] = None):
        self.config: WorkbookTemplateConfig = TemplateRegistry.load(profile)
        ensure_template_exists(self.config)
        self.workbook = openpyxl.load_workbook(self.config.template_path)

    def _sheet(self, category: str) -> Tuple[Worksheet, SheetMapping]:
        mapping = self.config.sheet_for(category)
        if mapping.sheet_name not in self.workbook.sheetnames:
            raise KeyError(f"Sheet '{mapping.sheet_name}' missing from loaded workbook template.")
        return self.workbook[mapping.sheet_name], mapping

    def write_metadata(self, category: str, field: str, value: Any):
        ws, mapping = self._sheet(category)
        cell_ref = mapping.metadata_cells.get(field)
        if not cell_ref:
            raise KeyError(f"No metadata cell configured for field '{field}' in category '{category}'.")
        ws[cell_ref] = value

    def append_row(self, category: str, row_data: Dict[str, Any]) -> int:
        """Writes one row of data into the next free row for this category. Returns the row index used."""
        ws, mapping = self._sheet(category)
        row_idx = self._find_next_free_row(ws, mapping)
        for field, value in row_data.items():
            col = mapping.field_to_column.get(field)
            if not col:
                continue
            cell = ws[f"{col}{row_idx}"]
            cell.value = value
            if isinstance(value, (int, float)):
                cell.number_format = '0.00"%"' if field.endswith("_pct") else "#,##0.00"
        return row_idx

    def flag_cell(self, category: str, row: int, field: str, message: str, severity: str = "WARNING"):
        """Injects a text-based alert into a specific cell/row and applies severity styling."""
        ws, mapping = self._sheet(category)
        col = mapping.field_to_column.get(field)
        if not col:
            raise KeyError(f"No column configured for field '{field}' in category '{category}'.")

        fill, font = SEVERITY_STYLES.get(severity, SEVERITY_STYLES["WARNING"])
        target_cell = ws[f"{col}{row}"]
        target_cell.fill = fill
        target_cell.border = THIN_BORDER

        notes_col = mapping.field_to_column.get("audit_notes")
        if notes_col:
            note_cell = ws[f"{notes_col}{row}"]
            note_cell.value = message
            note_cell.fill = fill
            note_cell.font = font
            note_cell.border = THIN_BORDER

    def stamp_last_synced(self, category: str):
        try:
            self.write_metadata(category, "last_synced_at", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"))
        except KeyError:
            pass

    def save(self, output_filename: str) -> str:
        output_dir = self.config.output_dir
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_filename)
        self.workbook.save(output_path)
        return output_path

    @staticmethod
    def _find_next_free_row(ws: Worksheet, mapping: SheetMapping) -> int:
        anchor_col = next(iter(mapping.field_to_column.values()))
        row = mapping.data_start_row
        while ws[f"{anchor_col}{row}"].value not in (None, ""):
            row += 1
        return row
