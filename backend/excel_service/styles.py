"""Reusable openpyxl style constants shared by the bootstrap builder and the writer."""
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

NAVY_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F497D")
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="595959")

OK_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

WARNING_FONT = Font(name="Calibri", size=11, bold=True, color="B78103")
CRITICAL_FONT = Font(name="Calibri", size=11, bold=True, color="C00000")

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

SEVERITY_STYLES = {
    "INFO": (OK_FILL, HEADER_FONT),
    "WARNING": (WARNING_FILL, WARNING_FONT),
    "CRITICAL": (CRITICAL_FILL, CRITICAL_FONT),
}
