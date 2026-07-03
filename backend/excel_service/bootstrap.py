"""
Builds a brand-new blank workbook from a WorkbookTemplateConfig if the
configured `template_path` doesn't exist yet. This is what today's
`build_lifecycle_sheets.init_master_system()` used to do by hand; here the
layout is derived entirely from the JSON profile so a new client profile
"just works" without touching this file.
"""
import os
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from backend.excel_service.config import WorkbookTemplateConfig
from backend.excel_service.styles import NAVY_FILL, HEADER_FONT, TITLE_FONT, LABEL_FONT

SHEET_TITLES = {
    "DPR": "DAILY PROGRESS EXTRACTION LEDGER",
    "MATERIAL": "MASTER MATERIAL & INVENTORY LEDGER",
    "BILLING": "VENDOR BILLING & MILESTONE TRACKER",
    "DRAWING": "EPC REGULATORY & DRAWING LOG",
}

METADATA_LABELS = {
    "last_synced_at": "Last Synced:",
    "report_date": "Report Date:",
    "log_category": "Log Category:",
    "labor_headcount": "Labor Headcount:",
    "structural_progress_pct": "Structural Progress %:",
}


def _humanize(field: str) -> str:
    return field.replace("_", " ").title()


def ensure_template_exists(config: WorkbookTemplateConfig):
    if os.path.exists(config.template_path):
        return
    build_template(config)


def build_template(config: WorkbookTemplateConfig):
    os.makedirs(os.path.dirname(config.template_path) or ".", exist_ok=True)
    os.makedirs(config.output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    default_sheet = wb.active
    is_first_sheet = True

    for category, mapping in config.sheets.items():
        ws = default_sheet if is_first_sheet else wb.create_sheet()
        ws.title = mapping.sheet_name
        is_first_sheet = False
        ws.sheet_view.showGridLines = True

        ws["A1"] = SHEET_TITLES.get(category, mapping.sheet_name.replace("_", " ").upper())
        ws["A1"].font = TITLE_FONT

        for field, cell_ref in mapping.metadata_cells.items():
            col_letters = "".join(filter(str.isalpha, cell_ref))
            row_digits = "".join(filter(str.isdigit, cell_ref))
            col_idx = column_index_from_string(col_letters)
            row_idx = int(row_digits)
            label_col = get_column_letter(max(col_idx - 1, 1))
            label_cell = ws[f"{label_col}{row_idx}"]
            label_cell.value = METADATA_LABELS.get(field, _humanize(field) + ":")
            label_cell.font = LABEL_FONT

        for field, col in mapping.field_to_column.items():
            header_cell = ws[f"{col}{mapping.header_row}"]
            header_cell.value = _humanize(field)
            header_cell.font = HEADER_FONT
            header_cell.fill = NAVY_FILL

        for col in mapping.field_to_column.values():
            ws.column_dimensions[col].width = 22

    wb.save(config.template_path)
    print(f"[excel_service] Bootstrapped new workbook template at {config.template_path}")
