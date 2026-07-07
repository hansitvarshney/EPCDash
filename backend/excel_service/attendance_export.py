"""
On-demand Site Attendance Sheet export.

Unlike the 4 ingestion categories (DPR/MATERIAL/BILLING/DRAWING), this isn't
routed through `ExcelWriterService`/`TemplateRegistry` -- there's no
"living" tracking template to append rows into here, just a fresh
diagnostic snapshot of the lifetime labor ledger generated on request. It
reuses the same corporate style palette from `styles.py` for visual
consistency with the rest of the platform's workbooks.
"""
from typing import Any, Dict, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from backend.excel_service.styles import HEADER_FONT, NAVY_FILL, THIN_BORDER, TITLE_FONT


def _write_header(ws: Worksheet, headers: List[str], row: int = 1) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _autosize(ws: Worksheet, widths: List[int]) -> None:
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


def build_attendance_workbook(
    site_name: str,
    daily_totals: List[Dict[str, Any]],
    ledger_rows: List[Dict[str, Any]],
) -> openpyxl.Workbook:
    """
    Builds a 2-sheet workbook:
      - "Daily Summary": one row per report_date with aggregate man-days.
      - "Raw Labor Ledger": one row per historical LaborLedger entry.
    Both sheets are chronological ascending (oldest first), matching the
    read-top-to-bottom convention of a physical logbook.
    """
    wb = openpyxl.Workbook()

    summary_ws = wb.active
    summary_ws.title = "Daily Summary"
    summary_ws["A1"] = f"{site_name} -- Attendance Daily Summary"
    summary_ws["A1"].font = TITLE_FONT
    summary_ws.merge_cells("A1:D1")

    _write_header(summary_ws, ["Report Date", "Masons", "Helpers", "Total Man-Days"], row=3)
    for idx, row in enumerate(daily_totals, start=4):
        summary_ws.cell(row=idx, column=1, value=row["report_date"])
        summary_ws.cell(row=idx, column=2, value=row["masons_count"])
        summary_ws.cell(row=idx, column=3, value=row["helpers_count"])
        summary_ws.cell(row=idx, column=4, value=row["total_man_days"])
        for col in range(1, 5):
            summary_ws.cell(row=idx, column=col).border = THIN_BORDER
    _autosize(summary_ws, [16, 12, 12, 16])
    summary_ws.freeze_panes = "A4"

    ledger_ws = wb.create_sheet("Raw Labor Ledger")
    _write_header(
        ledger_ws,
        ["Report Date", "Contractor", "Crew Type", "Masons", "Helpers", "Assigned Activity"],
        row=1,
    )
    for idx, row in enumerate(ledger_rows, start=2):
        ledger_ws.cell(row=idx, column=1, value=row["report_date"])
        ledger_ws.cell(row=idx, column=2, value=row["contractor_name"])
        ledger_ws.cell(row=idx, column=3, value=row["crew_type"])
        ledger_ws.cell(row=idx, column=4, value=row["masons_count"])
        ledger_ws.cell(row=idx, column=5, value=row["helpers_count"])
        ledger_ws.cell(row=idx, column=6, value=row["assigned_activity"])
        for col in range(1, 7):
            ledger_ws.cell(row=idx, column=col).border = THIN_BORDER
    _autosize(ledger_ws, [16, 32, 26, 10, 10, 36])

    return wb
