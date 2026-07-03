import os
import io
import openpyxl
import shutil
import re
import traceback
from pathlib import Path
from dotenv import load_dotenv
from datetime import datetime
from typing import List, Optional
from urllib.parse import unquote

# Calculate path to the root directory's .env file
backend_dir = Path(__file__).resolve().parent
root_dir = backend_dir.parent
env_path = root_dir / '.env'

# Load the root env file explicitly
load_dotenv(dotenv_path=env_path)

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware

from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from google import genai
from google.genai import types
from pydantic import BaseModel, Field, field_validator
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import time
from google.genai.errors import ServerError

# Absolute Package Imports to support executing from the workspace root
from backend.database import SessionLocal, Project, DailySiteLog, LaborLedger, DailyWorkMetrics, ProjectDocument
from backend.graph_query_engine import answer_project_query
from backend.analytics import AnalyticsEngine
from backend.schemas import AdaptiveLayoutConfig
from backend.graph_rag import build_project_knowledge_graph
from backend.build_lifecycle_sheets import process_daily_site_sheet

app = FastAPI(title="EPC Construction Audit & Document AI API")

origins = [
    "http://localhost:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Initialize Next-Gen Google GenAI client
client = genai.Client()

def normalize_extracted_date(raw_date_str: str) -> str:
    clean_str = raw_date_str.strip().replace("/", "-").replace(".", "-")
    formats = ("%Y-%m-%d", "%d-%m-%Y", "%y-%m-%d", "%d-%m-%y")
    for fmt in formats:
        try:
            return datetime.strptime(clean_str, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    print(f"⚠️ Warning: Could not match structural date format schema for string: '{raw_date_str}'")
    return clean_str

# ═════════════════════════════════════════════════════════
# 📋 STRUCTURED AUDITING SCHEMAS
# ═════════════════════════════════════════════════════════

class ExtractedQuantityEntry(BaseModel):
    work_category: str = Field(
        description=(
            "The human-readable classification of the work. Transform shorthand project symbols "
            "into clear, executive titles. For example: "
            "Convert 'SW' or 'SW-X' -> 'Shear Wall Shuttering Work', 'C' -> 'Column Shuttering'."
        )
    )
    element_id: str = Field(description="The precise structural tracking node code (e.g., 'SW-15', 'C-2', 'B-12').")
    raw_mathematical_dimensions: List[str] = Field(description="Extract each individual formula row string exactly as written.")
    claimed_subtotals_on_paper: List[float] = Field(
        default_factory=list,
        description=(
            "Extract the corresponding written subtotal number value for EACH individual formula row, "
            "ONLY when the paper actually writes one number per row. Leave this an empty list if the sheet "
            "instead shows a single consolidated 'Total Weight' figure covering multiple formula rows -- "
            "in that case populate 'element_final_total_weight' instead."
        )
    )
    element_final_total_weight: Optional[float] = Field(
        default=None,
        description=(
            "The single, clearly written closing 'Total' / 'Total Weight' figure for this element, if the "
            "sheet shows one consolidated total covering all of its formula rows (very common for reinforcement "
            "/ BBS steel weight calculations). Leave null ONLY when no such closing total line exists and every "
            "formula row already has its own individually written subtotal captured in 'claimed_subtotals_on_paper'."
        )
    )
    unit: str = Field(default="kg", description="The lowercase unit of measure.")

    @staticmethod
    def _coerce_numeric_string(v) -> Optional[float]:
        """Best-effort coercion of messy extracted number strings into a clean float."""
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        # Strip common unit/notation noise ('kg', 'nos', 'Kg.', commas, stray whitespace, diameter symbols)
        raw_str = str(v).lower()
        for token in ("kgs", "kg", "nos.", "nos", "φ"):
            raw_str = raw_str.replace(token, "")
        raw_str = raw_str.replace(",", "").strip().rstrip(".")
        if not raw_str:
            return None
        # Fall back to pulling the first numeric token out of the string (handles trailing junk)
        match = re.search(r"-?\d+(?:\.\d+)?", raw_str)
        if not match:
            return None
        try:
            return float(match.group())
        except ValueError:
            return None

    @field_validator("claimed_subtotals_on_paper", mode="before")
    @classmethod
    def sanitize_subtotals(cls, values):
        """Sanitizes text strings like '4202.547 kg' into pure floats before schema mapping."""
        if not isinstance(values, list):
            return values
        clean_values = []
        for v in values:
            coerced = cls._coerce_numeric_string(v)
            if coerced is None:
                print(f"⚠️ Could not coerce claimed subtotal value '{v}' into a float. Dropping (not zero-filling) so it doesn't silently dilute the element total.")
                continue
            clean_values.append(coerced)
        return clean_values

    @field_validator("element_final_total_weight", mode="before")
    @classmethod
    def sanitize_final_total(cls, value):
        """Sanitizes the single closing total figure (e.g. '4202.547 Kg', '4,202.55') into a pure float."""
        if value is None or value == "":
            return None
        coerced = cls._coerce_numeric_string(value)
        if coerced is None:
            print(f"⚠️ Could not coerce element_final_total_weight value '{value}' into a float. Ignoring.")
            return None
        return coerced

class ExtractedManpower(BaseModel):
    cumulative_masons: int = Field(default=0, description="Total active aggregated count of masons.")
    cumulative_helpers: int = Field(default=0, description="Total active aggregated count of helpers.")

class ProductionSiteLogPayload(BaseModel):
    report_date: str = Field(description="The actual log calendar date extracted from the header context. Strictly use 'YYYY-MM-DD'.")
    log_category: str = Field(default="GENERAL_CIVIL_WORKS")
    manpower_deployed: ExtractedManpower = Field(default_factory=ExtractedManpower)
    quantity_entries: List[ExtractedQuantityEntry]

# ═════════════════════════════════════════════════════════
# 🚀 MULTI-PAGE PRODUCTION STRUCTURED BATCH INGEST ROUTE
# ═════════════════════════════════════════════════════════

@app.post("/api/v1/fanout-ingest")
async def fanout_ingest(
    project_id: int, 
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db)
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        project = Project(id=project_id, name=f"Gurgaon Sector Layout Project {project_id}")
        db.add(project)
        db.commit()

    try:
        # --- 📦 Package all multi-page file assets into a sequential visual byte stream array ---
        visual_contents_list = []
        for file in files:
            image_bytes = await file.read()
            suffix = Path(file.filename).suffix.lower()
            mime_type = "application/pdf" if suffix == ".pdf" else "image/jpeg"
            visual_contents_list.append(types.Part.from_bytes(data=image_bytes, mime_type=mime_type))
        
        # --- Safe Retry wrapper for transient upstream 503 errors and 429 rate limits ---
        def execute_with_retry(prompt, config_obj):
            max_retries = 4
            backoff_delay = 5  
            for attempt in range(max_retries):
                try:
                    return client.models.generate_content(
                        model='gemini-2.5-flash',
                        contents=[
                            *visual_contents_list, 
                            prompt
                        ],
                        config=config_obj,
                    )
                except ServerError as se:
                    error_msg = str(se)
                    if ("503" in error_msg or "429" in error_msg or "Resource exhausted" in error_msg) and attempt < max_retries - 1:
                        print(f"⚠️ Gemini API rate limited or busy ({error_msg}). Retrying attempt {attempt + 1}/{max_retries} in {backoff_delay}s...")
                        time.sleep(backoff_delay)
                        backoff_delay *= 2  
                        continue
                    raise se

        # =====================================================================
        # SINGLE-PASS MULTI-PAGE STRUCTURING WITH ADVANCED MATHEMATICAL DIRECTIVES
        # =====================================================================
        unified_prompt = (
            "You are an expert construction data-compiler agent running an automated multi-page extraction pipeline.\n"
            "You are being provided with sequential files that make up a SINGLE daily site log report sheet.\n\n"
            
            "CRITICAL CONSOLIDATION DIRECTIVES:\n"
            "1. CONSOLIDATION CONTROL: Treat all provided images as part of the same continuous daily report ledger. Do not separate them into independent days.\n"
            "2. REPORT DATE: Locate the master report calendar date (typically on Page 1, top-right). Apply this single unified date to the root 'report_date' field.\n"
            "3. MANPOWER EXTRACTION (CUMULATIVE AGGREGATION):\n"
            "   - Read across all provided document pages. Accumulate the total workforce row counts sequentially.\n"
            "   - Ensure total summarized counts equal the aggregated sum total across all pages combined.\n"
            "4. QUANTITY SELECTION TRAVERSAL & MATH AGGREGATION:\n"
            "   - Track element tracking codes (e.g., SW-3, SW-8, SW-10) across the entire page collection.\n"
            "   - CASE A -- Per-row subtotals: If EVERY individual formula row already has its own written subtotal number beside it, "
            "put one value per row into 'claimed_subtotals_on_paper' (matching 'raw_mathematical_dimensions' 1-to-1), and leave 'element_final_total_weight' null.\n"
            "   - CASE B -- Single closing total (very common for reinforcement/BBS steel weight calculations with multiple bar-diameter/length/count rows like "
            "'20Φ-6m X 6', 'Lapping', 'Rings', 'Hooks'): if there is NO individual subtotal written next to each formula row, and instead only ONE consolidated "
            "'Total' / 'Total Weight' figure is written at the end covering ALL rows for that element, put that single number into 'element_final_total_weight' "
            "and leave 'claimed_subtotals_on_paper' as an empty list. Do NOT invent or split that single total across the rows.\n"
            "   - Never guess a value of 0 for a row or total you cannot actually read -- omit it instead.\n"
            "   - Explicitly strip text descriptors like 'kg', 'nos', or 'Nos' from numeric extractions so they validate cleanly as pure float numbers.\n"
            "   - Append all distinct item listings found into a single, comprehensive 'quantity_entries' JSON array sequentially."
        )

        print(f"⚡ Launching High-Speed Unified Pass for multi-page batch ({len(files)} files)...")
        start_time = time.time()

        structuring_config = types.GenerateContentConfig(
            response_mime_type="application/json",
            response_schema=ProductionSiteLogPayload,
            temperature=0.0,
        )
        
        final_json_response = execute_with_retry(unified_prompt, structuring_config)
        print(f"⏱️ Unified Multi-Page Extraction completed in {time.time() - start_time:.2f}s!")

        parsed_data = ProductionSiteLogPayload.model_validate_json(final_json_response.text)
        standard_date = normalize_extracted_date(parsed_data.report_date)

        # === 🛡️ DUPLICATE & RE-UPLOAD PREVENTION LAYER ===
        existing_log = db.query(DailySiteLog).filter(
            DailySiteLog.project_id == project_id,
            DailySiteLog.report_date == standard_date
        ).first()

        if existing_log:
            print(f"♻️ Existing log found for date {standard_date}. Purging old data cascades...")
            db.query(LaborLedger).filter(LaborLedger.log_id == existing_log.id).delete(synchronize_session='evaluate')
            db.query(DailyWorkMetrics).filter(DailyWorkMetrics.log_id == existing_log.id).delete(synchronize_session='evaluate')
            db.delete(existing_log)
            db.flush()
        
        site_log = DailySiteLog(
            project_id=project_id,
            report_date=standard_date,
            category=parsed_data.log_category
        )
        db.add(site_log)
        db.flush() 

        # Mount Excel Layout
        template_path = AdaptiveLayoutConfig.DEFAULT_TEMPLATE_PATH
        if not os.path.exists(template_path):
            raise HTTPException(status_code=500, detail=f"Master Excel template missing at path: {template_path}.")
            
        wb = openpyxl.load_workbook(template_path)
        
        amber_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        alert_font = Font(name="Calibri", size=11, bold=True, color="B78103")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        ws_log = wb["Daily_Progress_Log"]
        ws_log[AdaptiveLayoutConfig.METADATA_DATE_CELL] = standard_date
        ws_log[AdaptiveLayoutConfig.METADATA_CATEGORY_CELL] = parsed_data.log_category

        # --- RESOURCE ALLOCATION ROSTER ---
        total_masons = parsed_data.manpower_deployed.cumulative_masons or 0
        total_helpers = parsed_data.manpower_deployed.cumulative_helpers or 0
        calculated_labor_burn = total_masons + total_helpers

        ws_log[f"A{AdaptiveLayoutConfig.LABOR_ROW_START}"] = "Masons & Helpers Roster"
        ws_log[f"B{AdaptiveLayoutConfig.LABOR_ROW_START}"] = "Site Workforce Execution"
        ws_log[f"C{AdaptiveLayoutConfig.LABOR_ROW_START}"] = total_masons
        ws_log[f"D{AdaptiveLayoutConfig.LABOR_ROW_START}"] = total_helpers
        ws_log[f"E{AdaptiveLayoutConfig.LABOR_ROW_START}"] = f"Total Burn: {calculated_labor_burn} Man-Days"

        # 🛠️ FIXED: Removed 'project_id' parameter that triggered the invalid keyword argument exception
        db_labor = LaborLedger(
            log_id=site_log.id,
            contractor_name="General Structural Subcontractor",
            crew_type="Mixed Allocation Site Crew",
            masons_count=total_masons,
            helpers_count=total_helpers,
            assigned_activity="General Civil Works Execution"
        )
        db.add(db_labor)

        # --- PROGRESS ENTRIES WITH MULTI-PAGE MATH AUDITING ---
        q_row = AdaptiveLayoutConfig.QUANTITY_ROW_START
        resolved_weights_by_entry = []  # keeps the single source-of-truth totals for the overlay tabs below
        for q_entry in parsed_data.quantity_entries:
            calculated_total = 0.0
            is_mathematically_correct = True
            audit_notes = "Verified Clean"

            for formula, claimed_val in zip(q_entry.raw_mathematical_dimensions, q_entry.claimed_subtotals_on_paper):
                # Normalize handwritten multiplication notation ('×', capital/lowercase 'X') into valid
                # Python operators before attempting the arithmetic cross-check. Purely descriptive
                # engineering shorthand (e.g. rebar diameter 'Φ' notation) will still safely fail eval below.
                safe_formula = (
                    formula.replace("×", "*")
                    .replace("X", "*")
                    .replace("x", "*")
                    .replace(" ", "")
                )
                try:
                    true_line_value = eval(safe_formula, {"__builtins__": None}, {})
                    calculated_total += true_line_value
                    if abs(true_line_value - claimed_val) > 0.05:
                        is_mathematically_correct = False
                        audit_notes = f"Typo Found: Formula equals {true_line_value:.2f}, but paper claims {claimed_val}"
                except Exception:
                    continue

            claimed_subtotal_sum = sum(q_entry.claimed_subtotals_on_paper)

            # Priority order for the authoritative executed quantity:
            # 1) An explicit, clearly-written closing total (e.g. a single 'Total Weight' line covering
            #    several BBS/reinforcement formula rows that don't each carry their own subtotal).
            # 2) The sum of the individually written per-row subtotals.
            # 3) A recomputed fallback derived from evaluating the raw formulas, in case the paper
            #    genuinely provided neither of the above.
            if q_entry.element_final_total_weight:
                display_weight = q_entry.element_final_total_weight
            elif claimed_subtotal_sum:
                display_weight = claimed_subtotal_sum
            elif calculated_total:
                display_weight = calculated_total
            else:
                display_weight = 0.0
                print(f"⚠️ No usable quantity value found for element '{q_entry.element_id}' ({q_entry.work_category}). Storing 0.0 -- verify source document.")

            display_category = q_entry.work_category or parsed_data.log_category
            raw_element_id = q_entry.element_id
            
            ws_log[f"A{q_row}"] = display_category                  
            ws_log[f"B{q_row}"] = raw_element_id                    
            ws_log[f"C{q_row}"] = " + ".join(q_entry.raw_mathematical_dimensions)
            ws_log[f"D{q_row}"] = display_weight
            ws_log[f"E{q_row}"] = q_entry.unit
            
            db_metric = DailyWorkMetrics(
                log_id=site_log.id,
                category=display_category,                          
                element_id=raw_element_id,
                sub_component="General",
                formula_notation=" + ".join(q_entry.raw_mathematical_dimensions),
                metric_value=display_weight,
                unit=q_entry.unit
            )
            db.add(db_metric)
            resolved_weights_by_entry.append(display_weight)

            if not is_mathematically_correct:
                ws_log[f"D{q_row}"].fill = amber_fill
                ws_log[f"E{q_row}"] = audit_notes
                ws_log[f"F{q_row}"].fill = amber_fill
                ws_log[f"F{q_row}"].font = alert_font
                ws_log[f"F{q_row}"].alignment = Alignment(wrap_text=True)
                ws_log[f"F{q_row}"].border = thin_border
            q_row += 1

        # OVERLAYS -- reuse the already-resolved per-entry totals (element_final_total_weight ->
        # claimed subtotal sum -> recomputed fallback) instead of re-deriving from claimed_subtotals_on_paper
        # alone, which would silently zero out BBS/reinforcement entries that only carry a closing total.
        total_all_entries = sum(resolved_weights_by_entry)

        ws_inv = wb["Inventory_Reconciliation"]
        if parsed_data.log_category == "REINFORCEMENT_BBS":
            current_inv_value = ws_inv["D5"].value or 0
            ws_inv["D5"] = current_inv_value + total_all_entries

        ws_sch = wb["Project_Baseline_Tracker"]
        if parsed_data.log_category == "SHUTTERING" or "SHUTTERING" in str(parsed_data.log_category).upper():
            current_cumulative = ws_sch["D5"].value or 0
            ws_sch["D5"] = current_cumulative + total_all_entries

        db.flush()

        # ═════════════════════════════════════════════════════════
        # 📊 COMPILE MACRO EXECUTIVE ANALYTICS TAB
        # ═════════════════════════════════════════════════════════
        project_metrics = AnalyticsEngine.get_project_summary(db, project_id)
        
        if "Executive_Analytics" in wb.sheetnames:
            wb.remove(wb["Executive_Analytics"])
            
        ws_analytics = wb.create_sheet(title="Executive_Analytics")
        ws_analytics.views.sheetView[0].showGridLines = True
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        
        ws_analytics["A1"] = "PROJECT EXECUTIVE RUN-RUN CONTROL CENTER"
        ws_analytics["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        ws_analytics["A3"] = "Historical Resource Allocation Deployed"
        ws_analytics["A3"].font = bold_font
        
        ws_analytics.append(["Metric Type", "Total Accumulated Man-Days Deployed"])
        ws_analytics.append(["Skilled Force (Masons)", project_metrics["manpower_deployed"]["cumulative_masons"]])
        ws_analytics.append(["Helper Support Forces", project_metrics["manpower_deployed"]["cumulative_helpers"]])
        ws_analytics.append(["Total Project Labor Force Burn", project_metrics["manpower_deployed"]["total_man_days"]])
        
        for col in range(1, 3):
            ws_analytics.cell(row=7, column=col).fill = accent_fill
            ws_analytics.cell(row=7, column=col).font = bold_font
            
        ws_analytics["A9"] = "Cumulative Executed Quantities"
        ws_analytics["A9"].font = bold_font
        
        headers_qty = ["Work Category", "Structural Element Reference", "Total Output Executed To Date", "Unit of Measure"]
        ws_analytics.append(headers_qty)
        
        for col_idx, text in enumerate(headers_qty, start=1):
            cell = ws_analytics.cell(row=10, column=col_idx)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")

        for entry in project_metrics["quantities_executed"]:
            ws_analytics.append([entry["category"], entry["element_id"], entry["total_output"], entry["unit"]])
            
        for row in ws_analytics.iter_rows(min_row=4, max_row=ws_analytics.max_row, min_col=1, max_col=4):
            for cell in row:
                if cell.row not in [4, 10]:
                    cell.font = regular_font
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

        for col in ws_analytics.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_analytics.column_dimensions[col_letter].width = max(max_len + 3, 14)

        db.commit()
        
        output_filename = f"outputs/Master_Dashboard_Update_{project_id}_{standard_date}.xlsx"
        os.makedirs("outputs", exist_ok=True)
        wb.save(output_filename)
        return FileResponse(
            path=output_filename,
            filename=f"Master_Dashboard_Update_{standard_date}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        db.rollback()
        print("\n💥 --- PRODUCTION MULTI-PAGE STRUCTURED INGEST ENGINE ERROR ---")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Structured Batch Processing Fault: {str(e)}")

@app.post("/api/v1/projects/{project_id}/upload")
async def upload_project_document(project_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        project = Project(id=project_id, name=f"Project Workspace {project_id}")
        db.add(project)
        db.flush()

    storage_dir = f"backend/storage/project_{project_id}/documents"
    os.makedirs(storage_dir, exist_ok=True)
    local_path = os.path.join(storage_dir, file.filename)

    try:
        contents = await file.read()
        with open(local_path, "wb") as f:
            f.write(contents)
            
        if file.filename.lower().endswith(('.csv', '.xlsx', '.xls', '.jpg', '.jpeg', '.png')):
            try:
                success = process_daily_site_sheet(db=db, project_id=project_id, file_path=local_path)
            except Exception:
                traceback.print_exc()
                raise HTTPException(status_code=500, detail="Engine script error: (stripping_arr)")
            
            if success:
                return {"status": "success", "id": 9999, "file_name": file.filename, "message": "Site log successfully updated."}
            else:
                raise HTTPException(status_code=400, detail="Failed to parse standard structure layout.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to write file: {e}")

    exists = db.query(ProjectDocument).filter(ProjectDocument.project_id == project_id, ProjectDocument.file_name == file.filename).first()
    if not exists:
        file_category = "SPEC"
        if "contract" in file.filename.lower(): file_category = "CONTRACT"
        elif "tender" in file.filename.lower(): file_category = "TENDER"

        db_doc = ProjectDocument(project_id=project_id, file_name=file.filename, file_category=file_category, storage_path=local_path)
        db.add(db_doc)
        db.commit()

    try:
        build_project_knowledge_graph(project_id, db)
    except Exception as e:
        print(f"⚠️ Graph compilation deferred: {e}")

    return {"status": "success", "file_name": file.filename, "message": "File indexed successfully."}

class ChatQueryRequest(BaseModel):
    question: str

@app.post("/api/v1/projects/{project_id}/chat")
def chat_with_project_documents(project_id: int, payload: ChatQueryRequest):
    if not payload.question.strip(): raise HTTPException(status_code=400, detail="Question cannot be empty.")
    return {"status": "success", "project_id": project_id, "response": answer_project_query(project_id, payload.question)}

@app.get("/api/v1/analytics/summary/{project_id}")
def get_project_analytics_summary(project_id: int, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project: raise HTTPException(status_code=404, detail="Project profile metrics not found")
    return AnalyticsEngine.get_project_summary(db, project_id)

@app.get("/api/v1/analytics/velocity/{project_id}")
def get_project_velocity_timeline(project_id: int, days: int = 14, db: Session = Depends(get_db)):
    return AnalyticsEngine.get_velocity_trend(db, project_id, days=days)

@app.get("/projects")
def list_active_projects(db: Session = Depends(get_db)):
    projects_list = db.query(Project).all()
    if not projects_list:
        default_project = Project(name="Gurgaon Sector Project Workspace 1")
        db.add(default_project)
        db.commit()
        db.refresh(default_project)
        return [{"id": default_project.id, "name": default_project.name}]
    return [{"id": p.id, "name": p.name} for p in projects_list]

@app.delete("/projects/{project_id}/documents/{document_id}")
def delete_project_document(project_id: int, document_id: int, db: Session = Depends(get_db)):
    doc = db.query(ProjectDocument).filter(ProjectDocument.project_id == project_id, ProjectDocument.id == document_id).first()
    if not doc: raise HTTPException(status_code=404, detail="Document not found.")
    if os.path.exists(doc.storage_path): os.remove(doc.storage_path)
    index_path = f"backend/storage/project_{project_id}/doc_{document_id}"
    if os.path.exists(index_path): shutil.rmtree(index_path)
    db.delete(doc)
    db.commit()
    return {"status": "success", "message": f"Document {doc.file_name} removed successfully."}

@app.delete("/api/v1/projects/{project_id}/logs/{report_date}")
def delete_daily_site_log(project_id: int, report_date: str, db: Session = Depends(get_db)):
    decoded_date = unquote(report_date)
    standard_date = normalize_extracted_date(decoded_date)  
    site_log = db.query(DailySiteLog).filter(DailySiteLog.project_id == project_id, DailySiteLog.report_date == standard_date).first()
    if not site_log: raise HTTPException(status_code=404, detail=f"No site log data found for date {standard_date}")

    try:
        db.query(LaborLedger).filter(LaborLedger.log_id == site_log.id).delete(synchronize_session='evaluate')
        db.query(DailyWorkMetrics).filter(DailyWorkMetrics.log_id == site_log.id).delete(synchronize_session='evaluate')
        db.delete(site_log)
        db.commit()

        generated_excel_path = f"outputs/Master_Dashboard_Update_{project_id}_{standard_date}.xlsx"
        if os.path.exists(generated_excel_path): os.remove(generated_excel_path)
        return {"status": "success", "message": f"Log data for {standard_date} successfully erased."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database purge execution failed: {str(e)}")