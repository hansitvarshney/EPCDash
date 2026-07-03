import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from datetime import datetime
from sqlalchemy.orm import Session
from backend.database import DailySiteLog, LaborLedger, DailyWorkMetrics

def init_master_system():
    os.makedirs("templates", exist_ok=True)
    os.makedirs("outputs", exist_ok=True)
    wb = openpyxl.Workbook()
    
    # --- STYLE CONFIGS ---
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    green_fill = PatternFill(start_color="36648B", end_color="36648B", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    
    # SHEET 1: Daily Progress Log Ingestion
    ws1 = wb.active
    ws1.title = "Daily_Progress_Log"
    ws1.views.sheetView[0].showGridLines = True
    ws1["A1"] = "DAILY PROGRESS EXTRACTION LEDGER"
    ws1["A1"].font = title_font
    ws1["A3"] = "Report Date:"
    ws1["B3"] = "YYYY-MM-DD"
    
    headers_log = ["Element ID", "Sub-Tag", "Formula Notation", "Claimed Metric", "Unit", "System Audit Notes"]
    for i, h in enumerate(headers_log, 1):
        cell = ws1.cell(row=5, column=i, value=h)
        cell.font = white_bold; cell.fill = navy_fill; cell.alignment = Alignment(horizontal="center")

    # SHEET 2: Dynamic Material Inventory Balance Sheet
    ws2 = wb.create_sheet(title="Inventory_Reconciliation")
    ws2.views.sheetView[0].showGridLines = True
    ws2["A1"] = "LIVE SITE MATERIAL INVENTORY MONITOR"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="36648B")
    
    headers_inv = ["Material Description", "Unit", "Opening Balance", "Daily Quantity Used", "Current Remaining Stock"]
    for i, h in enumerate(headers_inv, 1):
        cell = ws2.cell(row=4, column=i, value=h)
        cell.font = white_bold; cell.fill = green_fill; cell.alignment = Alignment(horizontal="center")
        
    # Inject some mock starting inventory balances
    mock_stock = [
        ("20ø Reinforcement Steel", "Kg", 25000),
        ("25ø Reinforcement Steel", "Kg", 18000),
        ("Shuttering Ply Board Panels", "Nos", 1200),
    ]
    for idx, (mat, unit, bal) in enumerate(mock_stock, 5):
        ws2[f"A{idx}"] = mat
        ws2[f"B{idx}"] = unit
        ws2[f"C{idx}"] = bal
        ws2[f"D{idx}"] = 0 # Ingested dynamically
        ws2[f"E{idx}"] = f"=C{idx}-D{idx}" # Excel Formula native calculation!

    # SHEET 3: Schedule Remaining Work Target Tracker
    ws3 = wb.create_sheet(title="Project_Baseline_Tracker")
    ws3.views.sheetView[0].showGridLines = True
    ws3["A1"] = "MASTER SCHEDULE RUNNING TRACKER"
    ws3["A1"].font = title_font
    
    headers_sch = ["Work Activity Scope", "Unit", "Total Project Target (BOQ)", "Cumulative Done to Date", "Work Balance Remaining"]
    for i, h in enumerate(headers_sch, 1):
        cell = ws3.cell(row=4, column=i, value=h)
        cell.font = white_bold; cell.fill = navy_fill; cell.alignment = Alignment(horizontal="center")
        
    ws3["A5"] = "Shear Wall Shuttering Work"
    ws3["B5"] = "m2"
    ws3["C5"] = 5000.00  # Say this is his total project scope requirement
    ws3["D5"] = 1250.40  # Already executed in past days
    ws3["E5"] = "=C5-D5" # Excel native metric
    
    # Adjust widths
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 25

    wb.save("templates/master_site_template.xlsx")
    print("🚀 Blueprint Lifecycle System Initialized.")

if __name__ == "__main__":
    init_master_system()

def process_daily_site_sheet(db: Session, project_id: int, file_path: str, date_str: str = None):
    """
    Parses a daily site operational sheet (CSV/Excel) and updates 
    the SQLite labor ledger and work output metrics tables.
    """
    # 1. Parse or default the report logging date
    if not date_str:
        report_date = datetime.now().date()
    else:
        try:
            report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            report_date = datetime.now().date()

    # 2. Initialize or find the master DailySiteLog row for this date
    site_log = db.query(DailySiteLog).filter(
        DailySiteLog.project_id == project_id,
        DailySiteLog.report_date == report_date
    ).first()

    if not site_log:
        # 🛠️ FIX: Explicitly supply a default category to satisfy the SQLite NOT NULL constraint
        site_log = DailySiteLog(
            project_id=project_id, 
            report_date=report_date,
            category="General"  # 👈 This satisfies the database requirement!
        )
        db.add(site_log)
        db.commit()
        db.refresh(site_log)

    # 3. Read the uploaded operational layout sheet safely
    # 3. Read the uploaded operational layout sheet
    try:
        if file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
            
    except Exception as e:
        print(f"❌ Failed to parse data file format setup: {str(e)}")
        return False

    # 4. Aggregation Path A: Extract Headcounts (Masons, Helpers)
    # Assumes your columns match standard site rolls: 'role' and 'count' or similar
    total_masons = 0
    total_helpers = 0

    if 'role' in df.columns and 'count' in df.columns:
        mason_rows = df[df['role'].str.contains('mason', na=False, case=False)]
        helper_rows = df[df['role'].str.contains('helper|labor|labour', na=False, case=False)]
        
        total_masons = int(mason_rows['count'].sum())
        total_helpers = int(helper_rows['count'].sum())
    
    # Fallback/Direct structural match: if your columns are literally 'masons' and 'helpers'
    elif 'masons' in df.columns and 'helpers' in df.columns:
        total_masons = int(df['masons'].sum())
        total_helpers = int(df['helpers'].sum())

    # Update or insert Labor Ledger data records
    ledger_entry = db.query(LaborLedger).filter(LaborLedger.log_id == site_log.id).first()
    if not ledger_entry:
        ledger_entry = LaborLedger(log_id=site_log.id, masons_count=total_masons, helpers_count=total_helpers)
        db.add(ledger_entry)
    else:
        ledger_entry.masons_count += total_masons
        ledger_entry.helpers_count += total_helpers

    # 5. Aggregation Path B: Extract Material Quantities Executed
    # Assumes schema columns: 'category', 'element', 'quantity', 'unit'
    if all(col in df.columns for col in ['category', 'element', 'quantity']):
        for _, row in df.iterrows():
            metric_val = float(row['quantity']) if pd.notnull(row['quantity']) else 0.0
            if metric_val == 0:
                continue

            # 🛠️ FIX: Safe fallback to protect SQLite NOT NULL constraint
            raw_category = row.get('category')
            if pd.isna(raw_category) or str(raw_category).strip().lower() in ['none', '']:
                category_value = "General"
            else:
                category_value = str(raw_category).strip()

            metric_entry = DailyWorkMetrics(
                log_id=site_log.id,
                category=category_value,  # 👈 Clean safe variable inserted here
                element_id=str(row['element']).strip() if pd.notnull(row['element']) else "unmapped_node",
                metric_value=metric_val,
                unit=str(row.get('unit', 'cum')).strip() if pd.notnull(row.get('unit')) else "nos"
            )
            db.add(metric_entry)

    db.commit()
    return True   