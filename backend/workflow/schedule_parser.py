"""
Deterministic parser for uploaded Micro-Schedule master-schedule workbooks.

Unlike the other 4 ingestion categories, this never touches Gemini -- the
input is already a structured spreadsheet, so the extraction node parses it
directly with openpyxl. Header cells are matched by name (case-insensitive
aliases) rather than fixed column letters, so it tolerates the layout
variance of real-world site-engineer-maintained sheets.

A single sheet can carry two distinct row types, distinguished by an
optional "Type" column (defaults to PHYSICAL when absent, so every
pre-existing sheet with no such column keeps working unchanged):
  - PHYSICAL: pure site/engineering milestones (Substructure, Superstructure,
    MEP, ...) -- these populate `ProjectMilestone`.
  - PAYMENT: contract stage-payment tranches (RA Bills) that reference a
    PHYSICAL row by name via "Linked Milestone", plus a "Contract %" of the
    tender value -- these populate `PaymentMilestone` and are never mixed
    into the physical progress view.
"""
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional

import openpyxl

_HEADER_ALIASES: Dict[str, List[str]] = {
    "milestone_name": ["milestone", "milestone name", "phase", "phase name", "activity", "description", "task"],
    "target_date": ["target date", "date", "target", "due date", "planned date", "completion date", "end date"],
    "status": ["status", "state"],
    "sequence": ["sequence", "seq", "order", "#", "sl no", "sl. no.", "s.no", "s.no."],
    "type": ["type", "milestone type", "row type"],
    "linked_milestone": ["linked milestone", "linked physical milestone", "linked to", "linked stage"],
    "contract_pct": ["contract %", "contract pct", "% of contract value", "billing %", "payment %", "contract percentage"],
}

_REQUIRED_FIELDS = ("milestone_name", "target_date")

_DATE_FORMATS = ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d %b %Y")


def _parse_pct(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        # Values entered as a fraction (e.g. 0.15 via Excel's "%" cell format)
        # vs. a plain number (15) -- treat anything <= 1 as a fraction.
        return float(value) * 100 if value <= 1 else float(value)
    raw = str(value).strip().rstrip("%").strip()
    try:
        return float(raw)
    except ValueError:
        return 0.0


def _match_header(cell_value: Any) -> Optional[str]:
    if cell_value is None:
        return None
    normalized = str(cell_value).strip().lower()
    for field, aliases in _HEADER_ALIASES.items():
        if normalized in aliases:
            return field
    return None


def _parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def parse_milestone_workbook(file_bytes: bytes) -> Dict[str, List[Dict[str, Any]]]:
    """
    Returns {"physical": [...], "payments": [...]}.

    Physical rows: {milestone_name, target_date ("YYYY-MM-DD" or None),
    status ("PENDING"/"COMPLETED"), sequence (int)}.

    Payment rows: {bill_name, contract_pct (float), linked_physical_milestone_name
    (str or None), sequence (int)}.

    Raises ValueError with a clear message if the required Milestone/Target
    Date columns can't be located in the header row.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active

    header_map: Dict[str, int] = {}
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1), [])
    for cell in header_row:
        field = _match_header(cell.value)
        if field and field not in header_map:
            header_map[field] = cell.column - 1  # zero-indexed for values_only rows

    missing = [f for f in _REQUIRED_FIELDS if f not in header_map]
    if missing:
        raise ValueError(
            f"Could not find required column(s) {missing} in the uploaded schedule sheet. "
            f"Expected a header row with columns like 'Milestone Name' and 'Target Date'."
        )

    physical_rows: List[Dict[str, Any]] = []
    payment_rows: List[Dict[str, Any]] = []

    for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        name_col = header_map["milestone_name"]
        name = row[name_col] if name_col < len(row) else None
        if not name or not str(name).strip():
            continue
        name = str(name).strip()

        row_type = "PHYSICAL"
        if "type" in header_map:
            type_col = header_map["type"]
            raw_type = str(row[type_col]).strip().upper() if type_col < len(row) and row[type_col] else ""
            if raw_type in ("PAYMENT", "RA BILL", "BILLING", "STAGE PAYMENT"):
                row_type = "PAYMENT"

        sequence = idx
        if "sequence" in header_map:
            seq_col = header_map["sequence"]
            raw_seq = row[seq_col] if seq_col < len(row) else None
            if isinstance(raw_seq, (int, float)):
                sequence = int(raw_seq)

        if row_type == "PAYMENT":
            linked_name = None
            if "linked_milestone" in header_map:
                linked_col = header_map["linked_milestone"]
                raw_linked = row[linked_col] if linked_col < len(row) else None
                linked_name = str(raw_linked).strip() if raw_linked else None

            contract_pct = 0.0
            if "contract_pct" in header_map:
                pct_col = header_map["contract_pct"]
                contract_pct = _parse_pct(row[pct_col] if pct_col < len(row) else None)

            payment_rows.append(
                {
                    "bill_name": name,
                    "contract_pct": contract_pct,
                    "linked_physical_milestone_name": linked_name,
                    "sequence": sequence,
                }
            )
            continue

        date_col = header_map["target_date"]
        target_date = _parse_date(row[date_col]) if date_col < len(row) else None

        status = "PENDING"
        if "status" in header_map:
            status_col = header_map["status"]
            raw_status = str(row[status_col]).strip().upper() if status_col < len(row) and row[status_col] else ""
            if raw_status in ("PENDING", "COMPLETED"):
                status = raw_status
            elif raw_status in ("DONE", "COMPLETE", "FINISHED", "CLOSED"):
                status = "COMPLETED"

        physical_rows.append(
            {
                "milestone_name": name,
                "target_date": target_date.strftime("%Y-%m-%d") if target_date else None,
                "status": status,
                "sequence": sequence,
            }
        )

    return {"physical": physical_rows, "payments": payment_rows}
