import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def create_master_template():
    # Make sure templates directory exists at the root
    os.makedirs("templates", exist_ok=True)
    template_path = "templates/master_site_template.xlsx"
    
    wb = openpyxl.Workbook()
    
    # Styles Setup
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    # -------------------------------------------------------------
    # TAB 1: Daily_Progress_Log
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Daily_Progress_Log"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1["A1"] = "DAILY PROGRESS EXTRACTION MASTER LEDGER"
    ws1["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    
    # Metadata Placeholders
    ws1["A3"] = "Report Date:"
    ws1["A3"].font = bold_font
    ws1["B3"] = "YYYY-MM-DD"
    ws1["B3"].font = regular_font
    
    ws1["A4"] = "Active Phase Category:"
    ws1["A4"].font = bold_font
    ws1["B4"] = "PENDING"
    ws1["B4"].font = regular_font

    # Section A Headers: Resource/Labor Roster (Row 5)
    headers_labor = ["Contractor Name", "Crew Type / Trade", "Masons (Skilled)", "Helpers (Unskilled)", "Assigned Activity / Notes"]
    for col_idx, text in enumerate(headers_labor, start=1):
        cell = ws1.cell(row=5, column=col_idx, value=text)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    # Section B Headers: Structural Quantity Tracker (Row 17)
    headers_qty = ["Structural Element ID", "Sub-Component / Bar Tag", "Raw Formula Notation", "Claimed Measurement Value", "Unit", "Audit Flags / Verification Notes"]
    for col_idx, text in enumerate(headers_qty, start=1):
        cell = ws1.cell(row=17, column=col_idx, value=text)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    # -------------------------------------------------------------
    # TAB 2: Inventory_Reconciliation
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Inventory_Reconciliation")
    ws2.views.sheetView[0].showGridLines = True
    ws2["A1"] = "MATERIAL INVENTORY LOG & STOCKS FAN-OUT MATRIX"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="5BC0DE")
    
    headers_inv = ["Material Type", "Baseline Allocated", "Unit", "Cumulative Consumption To-Date"]
    for col_idx, text in enumerate(headers_inv, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=text)
        cell.fill = gray_fill
        cell.font = bold_font
    
    # Pre-populate Steel Row at Row 5
    ws2["A5"] = "Reinforcement Steel (ø BBS)"
    ws2["B5"] = 50000.0  # Placeholder baseline stock allocations
    ws2["C5"] = "kg"
    ws2["D5"] = 0.0     # Running aggregation cell target
    
    # -------------------------------------------------------------
    # TAB 3: Project_Baseline_Tracker
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Project_Baseline_Tracker")
    ws3.views.sheetView[0].showGridLines = True
    ws3["A1"] = "PROJECT SCHEDULE BASELINE TRANSITION INDEX"
    ws3["A1"].font = Font(name="Calibri", size=14, bold=True, color="5CB85C")
    
    headers_sch = ["WBS Task Phase Description", "Target Quantity Baseline", "Unit", "Audited Progress Quantity Matrix"]
    for col_idx, text in enumerate(headers_sch, start=1):
        cell = ws3.cell(row=4, column=col_idx, value=text)
        cell.fill = gray_fill
        cell.font = bold_font
        
    # Pre-populate Shuttering Row at Row 5
    ws3["A5"] = "Formwork / Shuttering Deployment Area"
    ws3["B5"] = 12500.0  # Target plan baseline
    ws3["C5"] = "m2"
    ws3["D5"] = 0.0     # Running aggregation cell target

    # Apply dimensions adjustments for layout readability
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 22)

    wb.save(template_path)
    print(f"SUCCESS: Created template layout asset successfully at: {template_path}")

if __name__ == "__main__":
    create_master_template()